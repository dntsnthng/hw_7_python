import zipfile
import csv
from openpyxl import load_workbook
from zipfile import ZipFile


def test_csv():
    with zipfile.ZipFile ('C:\\Users\\Admin\\Desktop\\getting-started-python-master\\qa_quru_homowork_7\\resourse\file.zip') as zip_file:
        with zip_file.open('file.csv', 'r') as csv_file: # открываем файл в архиве
            csv_text = io.TextIOWrapper(csv_file, encoding='utf-8')
            csvreader = csv.reader(csv_text)
            for row in csvreader:
                assert len(row) == 1



def test_pdf():
    with ZipFile("C:\\Users\\Admin\\Desktop\\getting-started-python-master\\qa_quru_homowork_7\\resourse\file.zip", 'r') as zip_file:
        with zip_file.open('file.pdf', 'r') as pdf_file:
            reader = Pdfreader (pdf_file)
            assert 'Работаем с файлами' in reader.pages[0].extract_text()





def test_xlsx(create_archive):
    with ZipFile("C:\\Users\\Admin\\Desktop\\getting-started-python-master\\qa_quru_homowork_7\\resourse\file.zip", 'r') as zip_file:
        with zip_file.open('file.xlsx') as xlsx_file:
            workbook = load_workbook('xlsx_file')
            sheet = workbook.active
            assert sheet.cell(row=4, column=2).value == 'sobaka@mail.ru'

